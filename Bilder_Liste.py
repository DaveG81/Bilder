#!/usr/bin/env python3

import os
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

# Pfad zum Hauptordner
main_folder = "/Users/david/Desktop/Bilder_Protokoll"

# Muster zur Erkennung der Agentur im Dateinamen
agency_pattern = re.compile(r"(Reuter|dpa|IMAGO|ddp|360°|action press|AFP|Getty|epa)", re.IGNORECASE)

# Dictionary, um für jede Agentur eine separate Workbook-Instanz zu speichern
agency_workbooks = {}

# Workbook für Dateien ohne Agenturnamen
no_agency_wb = None
no_agency_ws = None

# Standard Spaltenbreite festlegen
default_column_width = 10  # Openpyxl's default column width
adjusted_width_B_D = default_column_width * 4  # Viermal breiter für B und D
adjusted_width_A_C = default_column_width * 1.5  # 1,5 mal breiter für A und C

# Stil für die Schrift und Kopfzeile
font_style = Font(size=14)
header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Gelbe Füllung

# Funktion zur Anwendung des Stils auf die gesamte Tabelle
def style_worksheet(ws):
    for row in ws.iter_rows():
        for cell in row:
            cell.font = font_style
    for cell in ws[1]:  # Erste Zeile im Worksheet
        cell.fill = header_fill

# Durch alle Unterordner im Hauptordner iterieren
for root, dirs, files in os.walk(main_folder):
    for dir_name in dirs:
        # Unterordnername aufteilen mit ':' als Trennzeichen
        try:
            date, article = dir_name.split(':')
        except ValueError:
            print(f"Fehler beim Aufteilen des Ordnernamens: {dir_name}")
            continue

        # Pfad zum Unterordner
        subfolder_path = os.path.join(root, dir_name)

        # Durch alle Dateien im Unterordner iterieren
        for file_name in os.listdir(subfolder_path):
            if file_name.lower().endswith(('.png', '.jpeg', '.jpg')):
                # Agentur extrahieren (Spalte C)
                agency_match = agency_pattern.search(file_name)
                agency = agency_match.group(1) if agency_match else ""

                # Dateiname ohne Endung (Spalte D)
                photo_number = os.path.splitext(file_name)[0]

                # Wenn die Agentur erkannt wurde
                if agency:
                    # Prüfen, ob es bereits ein Workbook für diese Agentur gibt
                    if agency not in agency_workbooks:
                        # Neues Workbook für die Agentur erstellen
                        wb = Workbook()
                        ws = wb.active
                        ws.title = agency
                        ws.append(["DATUM", "ARTIKEL", "AGENTUR", "FOTO NR:"])
                        # Spaltenbreite anpassen
                        ws.column_dimensions['A'].width = adjusted_width_A_C
                        ws.column_dimensions['B'].width = adjusted_width_B_D
                        ws.column_dimensions['C'].width = adjusted_width_A_C
                        ws.column_dimensions['D'].width = adjusted_width_B_D
                        # Stil auf das gesamte Worksheet anwenden
                        style_worksheet(ws)
                        agency_workbooks[agency] = wb
                    else:
                        wb = agency_workbooks[agency]

                    ws = wb.active

                else:
                    # Wenn keine Agentur erkannt wurde, in die "Keine Agentur"-Datei schreiben
                    if no_agency_wb is None:
                        no_agency_wb = Workbook()
                        no_agency_ws = no_agency_wb.active
                        no_agency_ws.title = "Keine Agentur"
                        no_agency_ws.append(["DATUM", "ARTIKEL", "AGENTUR", "FOTO NR:"])
                        # Spaltenbreite anpassen
                        no_agency_ws.column_dimensions['A'].width = adjusted_width_A_C
                        no_agency_ws.column_dimensions['B'].width = adjusted_width_B_D
                        no_agency_ws.column_dimensions['C'].width = adjusted_width_A_C
                        no_agency_ws.column_dimensions['D'].width = adjusted_width_B_D
                        # Stil auf das gesamte Worksheet anwenden
                        style_worksheet(no_agency_ws)
                    
                    ws = no_agency_ws

                # Zeile in der Excel-Tabelle einfügen
                ws.append([date, article, agency, photo_number])

# Alle erstellten Excel-Dateien für Agenturen speichern
for agency, wb in agency_workbooks.items():
    output_path = f"/Users/david/Desktop/{agency}_Fotoabrechnung.xlsx"
    wb.save(output_path)
    print(f"Die Excel-Datei für {agency} wurde erfolgreich unter '{output_path}' gespeichert.")

# Excel-Datei für Dateien ohne Agenturnamen speichern, falls solche Dateien existieren
if no_agency_wb is not None:
    no_agency_output_path = "/Users/david/Desktop/Keine_Agentur_Fotoabrechnung.xlsx"
    no_agency_wb.save(no_agency_output_path)
    print(f"Die Excel-Datei für Dateien ohne Agentur wurde erfolgreich unter '{no_agency_output_path}' gespeichert.")
